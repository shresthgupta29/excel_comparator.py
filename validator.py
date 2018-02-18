from openpyxl import *
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.colors import COLOR_INDEX

wb1 = load_workbook(filename='C:/Users/shresgup/Downloads/sample1.xlsx')
wb2 = load_workbook(filename='C:/Users/shresgup/Downloads/sample2.xlsx')
dest_file = 'C:/Users/shresgup/Downloads/res.xlsx'
wb3 = Workbook()
ws1 = wb1.active
ws2 = wb2.active
ws3 = wb3.active
ws3.title = "result"
#my_red = colors.Color(COLOR_INDEX[2])
my_pass_fill = PatternFill(start_color=colors.COLOR_INDEX[3], end_color=colors.COLOR_INDEX[3],patternType='solid')

#my_red = colors.Color(COLOR_INDEX[3])
my_fail_fill = PatternFill(start_color=colors.COLOR_INDEX[2], end_color=colors.COLOR_INDEX[2],patternType='solid')

print(ws1.calculate_dimension())
print(ws2.calculate_dimension())
print(ws1.min_row)
print(ws1.max_column)
if (ws1.calculate_dimension() != ws2.calculate_dimension()):
    print("Dimension Mismatch")
    exit()
else:
    print("Same dimension")

for row in range(ws1.min_row, ws1.max_row+1):
    for col in range(ws1.min_column, ws1.max_column+1):
        if ws1.cell(row=row, column=col).value == ws2.cell(row=row, column=col).value:
            ws3.cell(row=row, column=col).value = 'Passed'
            ws3.cell(row=row, column=col).fill = my_pass_fill
        else:
            ws3.cell(row=row, column=col).value = 'Failed'
            ws3.cell(row=row, column=col).fill = my_fail_fill
wb3.save(filename=dest_file)
